from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

HEADERS = [
    'Butlovchi',
    'Belgi',
    'Miqdor',
    'Narx',
    'Summa',
    'Omborda',
    'Yetishmaydi',
    'Manba',
]


def build_configuration_workbook(configuration):
    """Configurator natijasini Excel chernovigiga aylantiradi."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Configurator'

    act_number = configuration.act.number if configuration.act else '-'
    sheet.append([f'Konfiguratsiya: {configuration.number}'])
    sheet.append([f'Bazaviy model: {configuration.base_product}'])
    sheet.append([f'Mijoz: {configuration.client or "-"}'])
    sheet.append([f'ACT: {act_number}'])
    sheet.append([f'Holat: {configuration.get_status_display()}'])
    sheet.append([])
    sheet.append(HEADERS)

    header_row = sheet.max_row
    for cell in sheet[header_row]:
        cell.font = Font(bold=True)

    for item in configuration.items.select_related('component'):
        sheet.append([
            item.component.name,
            item.label,
            item.quantity,
            float(item.unit_price),
            float(item.subtotal),
            float(item.available),
            float(item.shortage),
            'Ombordan' if item.source == 'stock' else 'Kirim qilinadi',
        ])

    sheet.append([])
    sheet.append(['', '', '', 'Jami:', float(configuration.total_price)])
    sheet.cell(row=sheet.max_row, column=4).font = Font(bold=True)
    sheet.cell(row=sheet.max_row, column=5).font = Font(bold=True)

    for index in range(1, len(HEADERS) + 1):
        sheet.column_dimensions[get_column_letter(index)].width = 18

    return workbook


def resolve_variant(configuration):
    """Konfiguratsiya uchun tayyor variantni topadi yoki yangisini yaratadi.

    TZ 6.2: bir xil tarkib avval bo'lgan bo'lsa — ombordagi tayyor pozitsiya
    va uning narxi ishlatiladi; bo'lmasa yangi variant omborga qo'shiladi va
    keyingi safar qayta ishlatiladi.
    """
    from django.db.transaction import atomic

    from apps.inventory.models import Product, ProductSpec

    existing = configuration.matching_variant
    if existing:
        return existing, False

    base = configuration.base_product
    with atomic():
        index = Product.objects.filter(base_model=base).count() + 1
        variant = Product.objects.create(
            sku=f'{base.sku}-V{index:02d}',
            name=f'{base.name} ({configuration.number})',
            kind=base.kind,
            description=f'{base.name} bazasida yig\'ilgan konfiguratsiya',
            cost_price=configuration.items_total,
            sale_price=configuration.items_total,
            base_model=base,
            signature=configuration.signature,
        )
        for item in configuration.items.select_related('component'):
            ProductSpec.objects.create(
                product=variant,
                component=item.component,
                label=item.label,
                quantity=item.quantity,
            )
    return variant, True


def finalize_modification(configuration, user, removal_overrides=None):
    """Tayyor mahsulotni o'zgartirishni yakunlaydi (modify rejimi, TZ 6.2).

    Ombor harakatlari:
      - butun bazaviy mahsulotdan 1 dona chiqim;
      - qo'shilgan butlovchilar ombordan chiqim;
      - yechib olinganlar omborga kirim (narxi bilan yozib qo'yiladi);
      - o'zgartirilgan mahsulot (variant) omborga 1 dona kirim.

    removal_overrides: {component_id: narx} — yechib olingan qism narxini
    o'zgartirish imkoniyati. Yakunda bugalterga xabar boradi.
    """
    from decimal import Decimal

    from django.db.transaction import atomic
    from rest_framework.exceptions import ValidationError

    from apps.core.models import Notification
    from apps.configurator.models import ConfigurationRemoval
    from apps.inventory.models import StockMovement
    from apps.inventory.services import apply_movement, available_quantity

    if not configuration.warehouse:
        raise ValidationError({'warehouse': "Tayyor mahsulotni o'zgartirish uchun ombor tanlanishi shart."})

    warehouse = configuration.warehouse
    base = configuration.base_product

    if available_quantity(base, warehouse) < 1:
        raise ValidationError({
            'detail': f'Omborda tayyor {base.name} qolmagan — o\'zgartirish uchun kamida 1 dona kerak.',
        })

    changes = configuration.changes
    shortages = [
        f"{row['component'].name} (kerak: {row['quantity']}, omborda: {available_quantity(row['component'], warehouse)})"
        for row in changes['added']
        if available_quantity(row['component'], warehouse) < row['quantity']
    ]
    if shortages:
        raise ValidationError({
            'detail': "Qo'shiladigan butlovchilar omborda yetarli emas.",
            'items': shortages,
        })

    overrides = {int(k): Decimal(str(v)) for k, v in (removal_overrides or {}).items()}

    with atomic():
        variant, created = resolve_variant(configuration)

        # 1 dona butun mahsulot ishga olinadi
        apply_movement(
            product=base, warehouse=warehouse,
            type=StockMovement.Type.OUT, quantity=1,
            reason=StockMovement.Reason.CONFIGURATION,
            reference=configuration.number, user=user,
        )
        # qo'shilgan butlovchilar ombordan
        for row in changes['added']:
            apply_movement(
                product=row['component'], warehouse=warehouse,
                type=StockMovement.Type.OUT, quantity=row['quantity'],
                reason=StockMovement.Reason.CONFIGURATION,
                reference=configuration.number, user=user,
            )
        # yechib olinganlar omborga qaytadi — narxi yozib qo'yiladi
        removed_lines = []
        for row in changes['removed']:
            price = overrides.get(row['component'].pk, row['unit_price'])
            ConfigurationRemoval.objects.create(
                configuration=configuration,
                component=row['component'],
                quantity=row['quantity'],
                unit_price=price,
                note='Tayyor mahsulotdan yechib olindi',
            )
            apply_movement(
                product=row['component'], warehouse=warehouse,
                type=StockMovement.Type.IN, quantity=row['quantity'],
                reason=StockMovement.Reason.CONFIGURATION,
                reference=configuration.number, user=user,
            )
            removed_lines.append(f"{row['component'].name} x{row['quantity']} — {price}")

        # o'zgartirilgan mahsulot tayyor pozitsiya sifatida omborga kiradi
        apply_movement(
            product=variant, warehouse=warehouse,
            type=StockMovement.Type.IN, quantity=1,
            reason=StockMovement.Reason.CONFIGURATION,
            reference=configuration.number, user=user,
        )

        # TZ: yechib olinganini ACT qilib bugalterga jo'natamiz
        act_number = configuration.act.number if configuration.act else '-'
        Notification.objects.create(
            title=f'{configuration.number}: tarkib o\'zgartirildi (ACT {act_number})',
            message=(
                'Yechib olindi va omborga qaytdi: ' + '; '.join(removed_lines)
                if removed_lines else 'Tarkibga faqat qo\'shimcha kiritildi.'
            ),
            level=Notification.Level.INFO,
            entity='Configuration',
            object_id=str(configuration.pk),
        )

    return variant, created


def copy_factory_spec(configuration):
    """Zavod tarkibini konfiguratsiya qatorlariga ko'chiradi (TZ 6.1).

    Model tanlanganda uning ichidagi barcha narsa tayyor keladi — keyin
    kerakli qatorlar o'zgartiriladi. Serializer ham, take_request ham
    aynan shu funksiyani chaqiradi (mantiq bitta joyda turadi).
    """
    from apps.configurator.models import ConfigurationItem

    for spec in configuration.base_product.specs.select_related('component'):
        ConfigurationItem.objects.create(
            configuration=configuration,
            component=spec.component,
            label=spec.label,
            quantity=spec.quantity,
        )


def take_request(request_obj, user, base_product=None, warehouse=None, mode=None):
    """Engineer zayavkani ishga oladi — chernovik konfiguratsiya avtomatik ochiladi.

    Bazaviy model: so'rov tanasidagi `base_product` > zayavkada yozilgani.
    Ikkalasi ham bo'lmasa 400 — konfiguratsiya modelsiz yaratilmaydi.
    """
    from django.db.transaction import atomic
    from rest_framework.exceptions import PermissionDenied, ValidationError

    from apps.configurator.models import Configuration, ConfigurationRequest
    from apps.inventory.models import Product

    if not (user.is_admin or user.is_engineer):
        raise PermissionDenied('Zayavkani faqat Engineer ishga oladi.')
    if request_obj.status != ConfigurationRequest.Status.NEW:
        raise ValidationError('Faqat yangi zayavkani ishga olish mumkin.')

    base_product = base_product or request_obj.base_product
    if base_product is None:
        raise ValidationError({
            'base_product': "Konfiguratsiya ochish uchun bazaviy model tanlanishi shart.",
        })
    if base_product.kind != Product.Kind.MACHINE:
        raise ValidationError({'base_product': 'Faqat tayyor model tanlanadi.'})

    with atomic():
        configuration = Configuration.objects.create(
            base_product=base_product,
            client=request_obj.client,
            warehouse=warehouse or request_obj.warehouse,
            mode=mode or Configuration.Mode.BUILD,
            note=f'{request_obj.number}: {request_obj.text}',
            created_by=user,
        )
        copy_factory_spec(configuration)

        request_obj.status = ConfigurationRequest.Status.IN_PROGRESS
        request_obj.taken_by = user
        request_obj.configuration = configuration
        request_obj.save()

    return request_obj


def complete_request(request_obj, user, configuration):
    """Engineer tayyor konfiguratsiyani biriktiradi — sales'ga xabar boradi."""
    from rest_framework.exceptions import PermissionDenied, ValidationError

    from apps.configurator.models import ConfigurationRequest
    from apps.core.models import Notification

    if not (user.is_admin or user.is_engineer):
        raise PermissionDenied('Zayavkani faqat Engineer yakunlaydi.')
    if request_obj.status not in {
        ConfigurationRequest.Status.NEW, ConfigurationRequest.Status.IN_PROGRESS,
    }:
        raise ValidationError('Zayavka allaqachon yakunlangan.')
    if configuration is None:
        raise ValidationError({'configuration': "Tayyor konfiguratsiya ko'rsatilishi shart."})

    request_obj.configuration = configuration
    request_obj.status = ConfigurationRequest.Status.DONE
    request_obj.taken_by = request_obj.taken_by or user
    request_obj.save()

    Notification.objects.create(
        user=request_obj.created_by,
        title=f'{request_obj.number}: konfiguratsiya tayyor',
        message=(
            f'{configuration.number} — {configuration.base_product.name}. '
            'Shartnoma jarayonini boshlashingiz mumkin.'
        ),
        level=Notification.Level.INFO,
        entity='ConfigurationRequest',
        object_id=str(request_obj.pk),
    )
    return request_obj


def notify_engineers_about_request(request_obj):
    """Yangi zayavka haqida barcha faol engineerlarga eslatma (3-xato tuzatmasi)."""
    from apps.accounts.models import User
    from apps.core.models import Notification

    engineers = User.objects.filter(role=User.Role.ENGINEER, is_active=True)
    for engineer in engineers:
        Notification.objects.create(
            user=engineer,
            title=f'{request_obj.number}: yangi zayavka',
            message=(request_obj.text or '')[:500],
            level=Notification.Level.INFO,
            entity='ConfigurationRequest',
            object_id=str(request_obj.pk),
        )
